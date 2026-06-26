import numpy as np
import math
from scipy import stats
import csv
import pandas as pd
import openpyxl
import copy

class GeneradorMultiplicadorConstante:
    def __init__(self, semilla_inicial, multiplicador):
        self.semilla_inicial = semilla_inicial
        self.semilla = semilla_inicial
        self.multiplicador = multiplicador
    
    def generar_numero_con_detalle(self, indice):
        # Guardamos la semilla actual (Xi)
        semilla_anterior = self.semilla
        
        # Y = X * a
        Y = self.semilla * self.multiplicador
        
        # Rellenar con ceros a la izquierda hasta 8 dígitos
        Y_str = str(Y).zfill(8)
        
        # Extraer los 4 dígitos centrales (índices 2, 3, 4, 5)
        self.semilla = int(Y_str[2:6])
        
        ri = self.semilla / 10000.0
        
        detalle = {
            "Indice": indice,
            "Semilla (Xi)": semilla_anterior,
            "Multiplicador (a)": self.multiplicador,
            "Producto (Xi * a)": Y,
            "Producto Relleno": Y_str,
            "Centro (Xi+1)": self.semilla,
            "Numero Pseudoaleatorio (Ri)": ri
        }
        
        return ri, detalle
    
    def generar_secuencia_completa(self, n):
        numeros = []
        detalles = []
        for i in range(n):
            ri, det = self.generar_numero_con_detalle(i + 1)
            numeros.append(ri)
            detalles.append(det)
        return numeros, detalles
    
    def generar_numero(self):
        # Y = X * a
        Y = self.semilla * self.multiplicador
        Y_str = str(Y).zfill(8)
        self.semilla = int(Y_str[2:6])
        return self.semilla / 10000.0
    
    def generar_secuencia(self, n):
        numeros = []
        for _ in range(n):
            numeros.append(self.generar_numero())
        return numeros
    
    def reset(self, nueva_semilla=None):
        self.semilla = nueva_semilla if nueva_semilla else self.semilla_inicial

def prueba_media(numeros, alpha=0.05):
    """Prueba de la media"""
    n = len(numeros)
    media_muestral = np.mean(numeros)
    media_teorica = 0.5
    varianza_teorica = 1/12
    
    z_alpha_2 = stats.norm.ppf(1 - alpha/2)
    error_estandar = math.sqrt(varianza_teorica / n)
    limite_inferior = media_teorica - z_alpha_2 * error_estandar
    limite_superior = media_teorica + z_alpha_2 * error_estandar
    
    aprobada = limite_inferior <= media_muestral <= limite_superior
    return aprobada, media_muestral, limite_inferior, limite_superior

def prueba_varianza(numeros, alpha=0.05):
    """Prueba de la varianza"""
    n = len(numeros)
    varianza_muestral = np.var(numeros, ddof=1)
    varianza_teorica = 1/12
    
    chi2_obs = (n - 1) * varianza_muestral / varianza_teorica
    chi2_inferior = stats.chi2.ppf(alpha/2, n-1)
    chi2_superior = stats.chi2.ppf(1-alpha/2, n-1)
    
    aprobada = chi2_inferior <= chi2_obs <= chi2_superior
    return aprobada, varianza_muestral, chi2_obs, chi2_inferior, chi2_superior

def prueba_chi_cuadrado(numeros, k=10, alpha=0.05):
    """Prueba de chi-cuadrado"""
    n = len(numeros)
    
    intervalos = np.linspace(0, 1, k+1)
    observadas = np.histogram(numeros, bins=intervalos)[0]
    esperadas = np.full(k, n/k)
    
    chi2_obs = np.sum((observadas - esperadas)**2 / esperadas)
    chi2_critico = stats.chi2.ppf(1-alpha, k-1)
    
    aprobada = chi2_obs <= chi2_critico
    return aprobada, chi2_obs, chi2_critico

def prueba_corridas(numeros, alpha=0.05):
    """Prueba de corridas arriba-abajo"""
    n = len(numeros)
    mediana = 0.5 # A veces se usa 0.5, a veces la mediana muestral
    simbolos = ['+' if x >= mediana else '-' for x in numeros]
    
    corridas = 1
    for i in range(1, len(simbolos)):
        if simbolos[i] != simbolos[i-1]:
            corridas += 1
    
    n1 = simbolos.count('+')
    n2 = simbolos.count('-')
    
    if n1 == 0 or n2 == 0:
        return False, 0, 0
    
    mu_r = (2 * n1 * n2) / (n1 + n2) + 1
    var_r = (2 * n1 * n2 * (2 * n1 * n2 - n1 - n2)) / ((n1 + n2)**2 * (n1 + n2 - 1))
    
    if var_r <= 0:
        return False, 0, 0
    
    z = (corridas - mu_r) / math.sqrt(var_r)
    z_critico = stats.norm.ppf(1 - alpha/2)
    
    aprobada = abs(z) <= z_critico
    return aprobada, z, z_critico

def copiar_estilo_celda(src_cell, dest_cell):
    if src_cell.has_style:
        dest_cell.font = copy.copy(src_cell.font)
        dest_cell.fill = copy.copy(src_cell.fill)
        dest_cell.border = copy.copy(src_cell.border)
        dest_cell.alignment = copy.copy(src_cell.alignment)
        dest_cell.number_format = src_cell.number_format

def generar_excel_desde_plantilla(template_path, output_path, is_demanda=True):
    try:
        wb = openpyxl.load_workbook(template_path, data_only=False)
        gen_sheet = wb.worksheets[0]
        sheet_name = gen_sheet.title
        
        if is_demanda:
            gen_sheet["G9"] = 5324.0
            gen_sheet["G10"] = 2026.0
            k_formula_template = '=IF(J{row}<=0.1287,0,IF(J{row}<=0.3926,1,IF(J{row}<=0.6631,2,IF(J{row}<=0.848,3,IF(J{row}<=0.9427,4,IF(J{row}<=0.9816,5,IF(J{row}<=0.9948,6,IF(J{row}<=0.9987,7,8))))))))'
        else:
            gen_sheet["G9"] = 4813.0
            gen_sheet["G10"] = 1754.0
            k_formula_template = '=MIN(14,INT(7+8*J{row}))'

        style_cells = {}
        for col_idx in range(1, 15): # columns A to N
            style_cells[col_idx] = gen_sheet.cell(row=14, column=col_idx)

        # Generar hasta la fila 10010 (10,000 números, del índice 0 al 9999)
        for i in range(3, 10000):
            row = i + 11
            
            cell_b = gen_sheet.cell(row=row, column=2, value=float(i))
            copiar_estilo_celda(style_cells[2], cell_b)
            
            cell_c = gen_sheet.cell(row=row, column=3, value=f'="Y"&B{row}&" = (X"&B{row}&")*a ="')
            copiar_estilo_celda(style_cells[3], cell_c)
            
            cell_d = gen_sheet.cell(row=row, column=4, value=f'=G{row-1}*$G$9')
            copiar_estilo_celda(style_cells[4], cell_d)
            
            cell_f = gen_sheet.cell(row=row, column=6, value=f'="X"&B{row}+1&" ="')
            copiar_estilo_celda(style_cells[6], cell_f)
            
            cell_g = gen_sheet.cell(row=row, column=7, value=f'=MID(TEXT(D{row}, "00000000"), 3, 4)')
            copiar_estilo_celda(style_cells[7], cell_g)
            
            cell_i = gen_sheet.cell(row=row, column=9, value=f'="R"&B{row}+1&" ="')
            copiar_estilo_celda(style_cells[9], cell_i)
            
            cell_j = gen_sheet.cell(row=row, column=10, value=f'=G{row}/10000')
            copiar_estilo_celda(style_cells[10], cell_j)
            
            cell_k = gen_sheet.cell(row=row, column=11, value=k_formula_template.format(row=row))
            copiar_estilo_celda(style_cells[11], cell_k)
            
            # Limpiar columnas de control (L, M, N)
            for c in [12, 13, 14]:
                gen_sheet.cell(row=row, column=c, value=None)

        # Primeros 20 números pseudoaleatorios
        for r_offset in range(10):
            gen_sheet.cell(row=12+r_offset, column=12, value=f"=J{11+r_offset}")
            gen_sheet.cell(row=12+r_offset, column=13, value=f"=J{21+r_offset}")

        # Prueba de Medias (Hoja 2)
        media_sheet = wb.worksheets[1]
        style_b = media_sheet.cell(row=4, column=2)
        for row in range(103, 10003):
            cell = media_sheet.cell(row=row, column=2, value=f"='{sheet_name}'!J{row+8}")
            copiar_estilo_celda(style_b, cell)
            
        media_sheet["E3"] = "=COUNT(B3:B10002)"
        media_sheet["E4"] = "=(SUM(B3:B10002))/$E$3"
        media_sheet["F4"] = "=AVERAGE(B3:B10002)"

        # Prueba de Varianza (Hoja 3)
        var_sheet = wb.worksheets[2]
        style_cells_var = {
            2: var_sheet.cell(row=4, column=2),
            3: var_sheet.cell(row=4, column=3),
            4: var_sheet.cell(row=4, column=4)
        }
        var_sheet.cell(row=103, column=4, value=None) # Limpiar sumatoria anterior
        
        for row in range(103, 10003):
            cell_b = var_sheet.cell(row=row, column=2, value=f"='{sheet_name}'!J{row+8}")
            copiar_estilo_celda(style_cells_var[2], cell_b)
            
            cell_c = var_sheet.cell(row=row, column=3, value=f"=B{row}-$G$4")
            copiar_estilo_celda(style_cells_var[3], cell_c)
            
            cell_d = var_sheet.cell(row=row, column=4, value=f"=POWER(C{row},2)")
            copiar_estilo_celda(style_cells_var[4], cell_d)

        # Nueva sumatoria
        var_sheet.cell(row=10003, column=4, value="=SUM(D3:D10002)")
        copiar_estilo_celda(style_cells_var[4], var_sheet.cell(row=10003, column=4))
        
        var_sheet["G3"] = "=COUNT(B3:B10002)"
        var_sheet["G8"] = "=D10003/($G$3-1)"
        var_sheet["H8"] = "=_xlfn.VAR.S(B3:B10002)"

        # Prueba Chi-Cuadrada (Hoja 4)
        chi_sheet = wb.worksheets[3]
        style_b_chi = chi_sheet.cell(row=5, column=2)
        for row in range(104, 10004):
            cell = chi_sheet.cell(row=row, column=2, value=f"='{sheet_name}'!J{row+7}")
            copiar_estilo_celda(style_b_chi, cell)
            
        chi_sheet["G5"] = "=COUNT(B4:B10003)"
        for r in range(10, 20):
            chi_sheet[f"I{r}"] = f'=COUNTIFS($B$4:$B$10003,">=" &G{r}, $B$4:$B$10003, "<=" &H{r})'

        # Pruebas Corridas (Hoja 5)
        corr_sheet = wb.worksheets[4]
        style_cells_corr = {
            2: corr_sheet.cell(row=5, column=2),
            4: corr_sheet.cell(row=5, column=4),
            5: corr_sheet.cell(row=5, column=5)
        }
        for row in range(103, 10003):
            cell_b = corr_sheet.cell(row=row, column=2, value=f"='{sheet_name}'!J{row+8}")
            copiar_estilo_celda(style_cells_corr[2], cell_b)
            
            cell_d = corr_sheet.cell(row=row, column=4, value=f"=IF(B{row}<=B{row-1},0,1)")
            copiar_estilo_celda(style_cells_corr[4], cell_d)
            
            cell_e = corr_sheet.cell(row=row, column=5, value=f'=IF(D{row}<>D{row-1},1,"")')
            copiar_estilo_celda(style_cells_corr[5], cell_e)

        corr_sheet["F4"] = "=SUM(E4:E10002)"
        corr_sheet["H10"] = "=COUNT(B3:B10002)"

        wb.save(output_path)
        print(f"Exportado a Excel exacto: {output_path}")
    except Exception as e:
        print(f"Error al exportar a Excel desde plantilla {template_path}: {e}")

def generar_y_evaluar(semilla, multiplicador, n_numeros, nombre_archivo_csv, nombre_archivo_excel, is_demanda):
    print(f"\n--- Generando {n_numeros} números para {nombre_archivo_csv} y {nombre_archivo_excel} ---")
    print(f"Semilla: {semilla}, Multiplicador: {multiplicador}")
    
    generador = GeneradorMultiplicadorConstante(semilla, multiplicador)
    numeros, detalles = generador.generar_secuencia_completa(n_numeros)
    
    print("\nResultados Pruebas Estadísticas:")
    
    # 1. Media
    aprobada_media, med, linf, lsup = prueba_media(numeros)
    print(f"[OK if aprobada_media else ERROR] Media: {med:.4f} (Intervalo: [{linf:.4f}, {lsup:.4f}])")
    
    # 2. Varianza
    aprobada_var, var, chi2_obs_var, chi2_inf_var, chi2_sup_var = prueba_varianza(numeros)
    print(f"[OK if aprobada_var else ERROR] Varianza: {var:.4f} (Chi2 Obs: {chi2_obs_var:.2f} en [{chi2_inf_var:.2f}, {chi2_sup_var:.2f}])")
    
    # 3. Chi Cuadrado
    k = 10
    aprobada_chi, chi2_obs_chi, chi2_critico_chi = prueba_chi_cuadrado(numeros, k=k)
    print(f"[OK if aprobada_chi else ERROR] Chi-Cuadrado: Obs = {chi2_obs_chi:.2f}, Crítico = {chi2_critico_chi:.2f}")
    
    # 4. Corridas
    aprobada_corridas, z_corridas, z_critico_corridas = prueba_corridas(numeros)
    print(f"[OK if aprobada_corridas else ERROR] Corridas: Z = {z_corridas:.4f}, Z_Crítico = {z_critico_corridas:.4f}")
    
    # Exportar CSV (para el simulador)
    with open(nombre_archivo_csv, 'w', newline='', encoding='utf-8') as file:
        writer = csv.writer(file)
        writer.writerow(["indice", "numero_pseudoaleatorio"])
        for i, numero in enumerate(numeros):
            writer.writerow([i+1, numero])
    print(f"Exportado a CSV: {nombre_archivo_csv}")
    
    # Exportar Excel usando plantilla
    template_path = "Generacion Numeros Pseudoaleatorios para Demanda.xlsx" if is_demanda else "Generacion Numeros Pseudoaleatorios  llegada del proveedor.xlsx"
    generar_excel_desde_plantilla(template_path, nombre_archivo_excel, is_demanda)
        
    return numeros

if __name__ == "__main__":
    print("Generador de Números Pseudoaleatorios - Multiplicador Constante (Entrega Pet)")
    print("=" * 70)
    
    CANTIDAD = 10000
    
    # Demanda: Semilla = 2026, Multiplicador = 5324
    nombre_excel_demanda = "Generacion Numeros Pseudoaleatorios para Demanda salida.xlsx"
    numeros_demanda = generar_y_evaluar(2026, 5324, CANTIDAD, "numeros_demanda.csv", nombre_excel_demanda, is_demanda=True)
    
    # Lead Time: Semilla = 1754, Multiplicador = 4813
    nombre_excel_lt = "Generacion Numeros Pseudoaleatorios  llegada del proveedor salida.xlsx"
    numeros_lead_time = generar_y_evaluar(1754, 4813, CANTIDAD, "numeros_lead_time.csv", nombre_excel_lt, is_demanda=False)
    
    print("\nProceso finalizado.")
