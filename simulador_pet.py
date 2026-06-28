import scipy.stats as stats
import math
import csv
import json
import os
import shutil
import subprocess
import sys
from datetime import datetime
from pathlib import Path
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

class SimuladorInventario:
    def __init__(self, archivo_demanda, archivo_lead_time):
        self.numeros_demanda = self.cargar_csv(archivo_demanda)
        self.numeros_lead_time = self.cargar_csv(archivo_lead_time)
        self.demandas_transformadas = [
            int(valor) for valor in stats.poisson.ppf(self.numeros_demanda, 2.05)
        ]
        self.idx_demanda = 0
        self.idx_lead_time = 0
        
        # Variables exógenas de costos
        # $2.500.000 / (300 bolsas * 30 días) = $277,78 por bolsa/día
        self.CALM = round(2_500_000 / (300 * 30), 2)
        self.CVP = 10430     # $52.150 - $41.720 por cada venta perdida
        # $2.000.000 / (24.000 kg / 20 kg por bolsa) = $1.666,67 por bolsa
        self.CEP = round(2_000_000 / (24_000 / 20), 2)

    def cargar_csv(self, archivo):
        numeros = []
        try:
            with open(archivo, 'r', encoding='utf-8') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    numeros.append(float(row['numero_pseudoaleatorio']))
        except Exception as e:
            print(f"Error al cargar {archivo}: {e}")
        return numeros

    def reiniciar_secuencias(self):
        """Reinicia los índices para comparar políticas con los mismos aleatorios."""
        self.idx_demanda = 0
        self.idx_lead_time = 0

    def obtener_ri_demanda(self):
        ri = self.numeros_demanda[self.idx_demanda]
        self.idx_demanda = (self.idx_demanda + 1) % len(self.numeros_demanda)
        return ri

    def obtener_ri_lead_time(self):
        ri = self.numeros_lead_time[self.idx_lead_time]
        self.idx_lead_time = (self.idx_lead_time + 1) % len(self.numeros_lead_time)
        return ri

    def generar_demanda(self):
        indice = self.idx_demanda
        ri = self.numeros_demanda[indice]
        demanda = self.demandas_transformadas[indice]
        self.idx_demanda = (indice + 1) % len(self.numeros_demanda)
        return demanda, ri

    def generar_lead_time(self):
        ri = self.obtener_ri_lead_time()
        # Transformada Inversa Uniforme Discreta [7, 14]
        return math.floor(7 + (8 * ri)), ri

    def simular_politica(self, PEP, TP, TF=180, validacion_manual=False):
        # CONDICIONES INICIALES SEGÚN EL DIAGRAMA
        T = 0
        FLL = 1
        ST = 0
        PP = True
        CTALM = 0
        CVTAP = 0
        CTEP = 0
        
        # Métricas para resumen
        demanda_total = 0
        ventas_perdidas_total = 0
        pedidos_realizados = 0
        
        datos_validacion = []
        numeros_usados_dem = []
        numeros_usados_lt = []
        
        # Bucle de incremento de tiempo constante
        for _ in range(TF):
            T += 1
            llegada_pedido = False

            # El Excel muestra el pedido que estaba pendiente al comenzar el día.
            pedido_pendiente_inicio = PP
            fll_inicio = FLL if PP else None
            
            # 1. ¿T = FLL? (Llegada de pedido)
            if PP and T == FLL:
                ST += TP
                llegada_pedido = True
                PP = False

            # Generar Demanda
            DD, ri_dem = self.generar_demanda()
            numeros_usados_dem.append(ri_dem)
            demanda_total += DD
            
            st_anterior = ST
            vp_dia = 0
            costo_almacenamiento_dia = 0
            costo_ventas_perdidas_dia = 0
            # 2. ¿ST >= DD?
            if ST >= DD:
                ST -= DD
                costo_almacenamiento_dia = ST * self.CALM
                CTALM += costo_almacenamiento_dia
            else:
                vp_dia = DD - ST
                costo_ventas_perdidas_dia = vp_dia * self.CVP
                CVTAP += costo_ventas_perdidas_dia
                ST = 0
                
            ventas_perdidas_total += vp_dia
                
            # 3. Revisión de inventario
            emite_pedido = False
            lt_dia = None
            ri_lt = None
            costo_emision_dia = 0
            if ST <= PEP and not PP:
                lt_dia, ri_lt = self.generar_lead_time()
                numeros_usados_lt.append(ri_lt)
                FLL = T + lt_dia
                costo_emision_dia = self.CEP * TP
                CTEP += costo_emision_dia
                emite_pedido = True
                pedidos_realizados += 1
                PP = True

            if validacion_manual:
                costo_total_dia = costo_emision_dia + costo_almacenamiento_dia + costo_ventas_perdidas_dia
                datos_validacion.append({
                    "Fila": T,
                    "Dia": T,
                    "FLL": fll_inicio,
                    "Ri_Lead_Time": ri_lt,
                    "Tamaño_Pedido": TP if pedido_pendiente_inicio else None,
                    "Ri_Demanda": ri_dem,
                    "Stock_Inicial": st_anterior,
                    "Llega_Pedido": "SI" if llegada_pedido else "NO",
                    "Demanda": DD,
                    "Ventas": DD - vp_dia,
                    "Stock_Final": ST,
                    "PEP": PEP,
                    "Pedido_Pendiente": "SI" if pedido_pendiente_inicio else "NO",
                    "Ventas_Perdidas": vp_dia,
                    "Emite_Pedido": "SI" if emite_pedido else "NO",
                    "Lead_Time": lt_dia,
                    "Costo_Emision_Dia": costo_emision_dia,
                    "Costo_Almacenamiento_Dia": costo_almacenamiento_dia,
                    "Costo_Ventas_Perdidas_Dia": costo_ventas_perdidas_dia,
                    "Costo_Total_Dia": costo_total_dia,
                    "CTALM_Acumulado": CTALM,
                    "CVTAP_Acumulado": CVTAP,
                    "CTEP_Acumulado": CTEP,
                    "CTF_Acumulado": CTALM + CVTAP + CTEP
                })

        # Resultados de la corrida
        CTF = CTALM + CVTAP + CTEP
        
        if validacion_manual:
            return {
                "PEP": PEP,
                "TP": TP,
                "CALM": self.CALM,
                "CVP": self.CVP,
                "CEP": self.CEP,
                "CTF": CTF,
                "Demanda_Total": demanda_total,
                "Ventas_Perdidas": ventas_perdidas_total,
                "Pedidos_Realizados": pedidos_realizados,
                "CTALM": CTALM,
                "CVTAP": CVTAP,
                "CTEP": CTEP,
                "Datos_Validacion": datos_validacion,
                "Numeros_Usados_Dem": numeros_usados_dem,
                "Numeros_Usados_LT": numeros_usados_lt
            }
            
        return CTF

def dar_formato_excel(filepath):
    try:
        wb = openpyxl.load_workbook(filepath)
        sheet = wb.active
        
        # Estilos visuales premium (Azul oscuro elegante)
        header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
        header_font = Font(name="Segoe UI", size=11, bold=True, color="FFFFFF")
        data_font = Font(name="Segoe UI", size=10)
        align_center = Alignment(horizontal="center", vertical="center")
        thin_side = Side(border_style="thin", color="D3D3D3")
        border_all = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
        
        # Dar formato a la cabecera
        for col in range(1, sheet.max_column + 1):
            cell = sheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = align_center
            cell.border = border_all
            
        # Dar formato a los datos de la tabla
        for row in range(2, sheet.max_row + 1):
            for col in range(1, sheet.max_column + 1):
                cell = sheet.cell(row=row, column=col)
                cell.font = data_font
                cell.alignment = align_center
                cell.border = border_all
                
                # Detectar columnas con dinero acumulado y aplicar formato moneda en pesos sin centavos
                header_name = sheet.cell(row=1, column=col).value
                if header_name and "Acumulado" in header_name:
                    cell.number_format = "$#,##0"
                elif header_name and header_name in ["Dia", "Stock_Inicial", "Demanda", "Stock_Final", "Ventas_Perdidas", "Lead_Time"]:
                    cell.number_format = "#,##0"
                    
        # Autoajuste de ancho de columnas
        for col in sheet.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = openpyxl.utils.get_column_letter(col[0].column)
            sheet.column_dimensions[col_letter].width = max(max_len + 4, 12)
            
        wb.save(filepath)
    except Exception as e:
        print(f"Error al dar formato al Excel {filepath}: {e}")

def crear_carpeta_corrida(carpeta_base=os.path.join("salidas", "corridas")):
    """Crea una carpeta única para no sobrescribir ejecuciones anteriores."""
    identificador = datetime.now().strftime("%Y-%m-%d_%H-%M-%S_%f")
    carpeta = os.path.abspath(os.path.join(carpeta_base, f"corrida_{identificador}"))
    os.makedirs(carpeta, exist_ok=False)
    return carpeta

def exportar_simulacion_completa(
    resumen,
    resultados_politicas=None,
    mejor_politica=None,
    filepath="Simulación de inventario.xlsx",
    mostrar_mensaje=True
):
    """Exporta una corrida diaria con una estructura equivalente al Excel de referencia."""
    if not resumen["Datos_Validacion"]:
        return

    columnas = [
        ("Fila", "FILA"),
        ("Dia", "T (día actual)"),
        ("FLL", "FLL (fecha de llegada del pedido)"),
        ("Ri_Lead_Time", "Número aleatorio del tiempo de entrega"),
        ("Lead_Time", "Tiempo de entrega del proveedor"),
        ("Tamaño_Pedido", "Tamaño del pedido"),
        ("Ri_Demanda", "Número aleatorio de la demanda"),
        ("Demanda", "Demanda diaria"),
        ("Stock_Inicial", "ST INICIAL"),
        ("Demanda", "DEMANDA"),
        ("Ventas", "VENTAS"),
        ("Stock_Final", "ST FINAL"),
        ("PEP", "PEP (punto de emisión del pedido)"),
        ("Pedido_Pendiente", "PEDIDO PENDIENTE (PP)"),
        ("Ventas_Perdidas", "Ventas perdidas"),
        ("Costo_Emision_Dia", "Costo emisión pedido"),
        ("Costo_Almacenamiento_Dia", "Costo almacenamiento"),
        ("Costo_Ventas_Perdidas_Dia", "Costo ventas perdidas"),
        ("Costo_Total_Dia", "Costo total"),
    ]

    filas = []
    for registro in resumen["Datos_Validacion"]:
        fila = {}
        for clave, encabezado in columnas:
            valor = registro.get(clave)
            fila[encabezado] = None if valor == "-" else valor
        filas.append(fila)

    pd.DataFrame(filas).to_excel(filepath, index=False, sheet_name="Simulación diaria")

    wb = openpyxl.load_workbook(filepath)
    ws = wb["Simulación diaria"]
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:S{ws.max_row}"
    ws.sheet_view.showGridLines = False
    ws.row_dimensions[1].height = 74

    encabezado_general = PatternFill("solid", fgColor="FFFFFF")
    encabezado_pendiente = PatternFill("solid", fgColor="FFFF00")
    encabezado_costos = PatternFill("solid", fgColor="C5E0B3")
    relleno_aleatorio_entrega = PatternFill("solid", fgColor="D9EAD3")
    relleno_blanco = PatternFill("solid", fgColor="FFFFFF")
    fuente_encabezado = Font(name="Calibri", size=11, bold=False, color="000000")
    fuente_costos = Font(name="Calibri", size=11, bold=True, color="2F5496")
    borde = Border(
        left=Side(style="thin", color="000000"),
        right=Side(style="thin", color="000000"),
        top=Side(style="thin", color="000000"),
        bottom=Side(style="thin", color="000000"),
    )

    for celda in ws[1]:
        celda.fill = encabezado_general
        celda.font = fuente_encabezado
        celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        celda.border = borde

    ws["N1"].fill = encabezado_pendiente
    for col in range(16, 20):
        ws.cell(1, col).fill = encabezado_costos
        ws.cell(1, col).font = fuente_costos
        ws.cell(1, col).alignment = Alignment(horizontal="left", vertical="center")

    for row in range(2, ws.max_row + 1):
        ws.row_dimensions[row].height = 15
        for col in range(1, 20):
            celda = ws.cell(row, col)
            celda.border = borde
            celda.fill = relleno_blanco
            celda.font = Font(name="Calibri", size=11, color="000000")
            celda.alignment = Alignment(horizontal="right", vertical="center")

        for col in [1, 2, 3, 5, 6, 8, 14]:
            ws.cell(row, col).alignment = Alignment(horizontal="center", vertical="center")

        for col in [4, 5, 7, 8]:
            ws.cell(row, col).font = Font(name="Arial", size=10, color="000000")

        ws.cell(row, 4).fill = relleno_aleatorio_entrega
        for col in range(16, 20):
            ws.cell(row, col).number_format = '#,##0'
        ws.cell(row, 4).number_format = "0.0000"
        ws.cell(row, 7).number_format = "0.0000"

    anchos = {
        "A": 4.71, "B": 7.00, "C": 12.14, "D": 10.14, "E": 13.43,
        "F": 10.43, "G": 9.14, "H": 8.71, "I": 9.71, "J": 10.29,
        "K": 7.71, "L": 8.57, "M": 16.00, "N": 10.86, "O": 8.43,
        "P": 19.86, "Q": 21.14, "R": 20.43, "S": 10.71,
    }
    for columna, ancho in anchos.items():
        ws.column_dimensions[columna].width = ancho

    if resultados_politicas and mejor_politica:
        resumen_ws = wb.create_sheet("Resumen políticas", 0)
        resumen_ws.sheet_view.showGridLines = False
        resumen_ws.freeze_panes = "A7"

        resumen_ws.merge_cells("A1:I1")
        resumen_ws["A1"] = "Comparación de políticas de inventario"
        resumen_ws["A1"].fill = PatternFill("solid", fgColor="1F4E79")
        resumen_ws["A1"].font = Font(name="Calibri", size=16, bold=True, color="FFFFFF")
        resumen_ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
        resumen_ws.row_dimensions[1].height = 30

        resumen_ws.merge_cells("A2:I2")
        resumen_ws["A2"] = (
            "Cada combinación PEP/TP corresponde a una única corrida de 180 días."
        )
        resumen_ws["A2"].alignment = Alignment(horizontal="center")
        resumen_ws["A2"].font = Font(name="Calibri", size=10, italic=True, color="595959")

        resumen_ws.merge_cells("A3:I3")
        resumen_ws["A3"] = (
            f"Costos vigentes: CALM ${resumen['CALM']:,}/unidad-día · "
            f"CVP ${resumen['CVP']:,}/venta perdida · "
            f"CEP ${resumen['CEP']:,}/bolsa pedida"
        )
        resumen_ws["A3"].alignment = Alignment(horizontal="center")
        resumen_ws["A3"].font = Font(name="Calibri", size=10, color="595959")

        mejor_resultado = min(resultados_politicas, key=lambda x: x["CTF"])
        resumen_ws["A4"] = "Política óptima"
        resumen_ws["B4"] = f"PEP {mejor_politica[0]} / TP {mejor_politica[1]}"
        resumen_ws["D4"] = "CTF mínimo"
        resumen_ws["E4"] = mejor_resultado["CTF"]
        resumen_ws["G4"] = "Corridas realizadas"
        resumen_ws["H4"] = len(resultados_politicas)

        for celda in ["A4", "D4", "G4"]:
            resumen_ws[celda].fill = PatternFill("solid", fgColor="D9EAF7")
            resumen_ws[celda].font = Font(bold=True, color="1F1F1F")
            resumen_ws[celda].alignment = Alignment(horizontal="center")
        for celda in ["B4", "E4", "H4"]:
            resumen_ws[celda].fill = PatternFill("solid", fgColor="E2F0D9")
            resumen_ws[celda].font = Font(bold=True, color="006100")
            resumen_ws[celda].alignment = Alignment(horizontal="center")
        resumen_ws["E4"].number_format = '$#,##0'

        encabezados = [
            "Ranking", "PEP", "TP", "Costo almacenamiento",
            "Costo ventas perdidas", "Costo pedidos", "CTF",
            "Nivel de servicio", "Conclusión"
        ]
        for col, encabezado in enumerate(encabezados, start=1):
            celda = resumen_ws.cell(6, col, encabezado)
            celda.fill = PatternFill("solid", fgColor="C6E0B4")
            celda.font = Font(bold=True, color="2F5496")
            celda.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            celda.border = borde

        resultados_ordenados = sorted(resultados_politicas, key=lambda x: x["CTF"])
        for ranking, resultado in enumerate(resultados_ordenados, start=1):
            fila = ranking + 6
            es_optima = (resultado["PEP"], resultado["TP"]) == mejor_politica
            es_primera = (resultado["PEP"], resultado["TP"]) == (5, 20)
            conclusion = "ÓPTIMA" if es_optima else ("Primera alternativa" if es_primera else "")
            valores = [
                ranking, resultado["PEP"], resultado["TP"], resultado["CTALM"],
                resultado["CVTAP"], resultado["CTEP"], resultado["CTF"],
                resultado["Nivel_Servicio"], conclusion
            ]
            for col, valor in enumerate(valores, start=1):
                celda = resumen_ws.cell(fila, col, valor)
                celda.border = borde
                celda.alignment = Alignment(horizontal="center", vertical="center")
                if col in [4, 5, 6, 7]:
                    celda.number_format = '$#,##0'
                elif col == 8:
                    celda.number_format = '0.0%'
                if es_optima:
                    celda.fill = PatternFill("solid", fgColor="C6EFCE")
                    celda.font = Font(bold=True, color="006100")
                elif es_primera:
                    celda.fill = PatternFill("solid", fgColor="D9EAF7")

        ultima_fila = 6 + len(resultados_ordenados)
        resumen_ws.auto_filter.ref = f"A6:I{ultima_fila}"
        for columna, ancho in {
            "A": 19, "B": 15, "C": 10, "D": 21, "E": 22,
            "F": 17, "G": 17, "H": 18, "I": 22
        }.items():
            resumen_ws.column_dimensions[columna].width = ancho

        wb.active = 0

    wb.save(filepath)
    if mostrar_mensaje:
        print(f"Exportado a Excel de simulación completa: {filepath}")

def exportar_resumen_validacion(
    resumen,
    resultados_politicas=None,
    mejor_politica=None,
    carpeta_salida="."
):
    archivo_csv = os.path.join(carpeta_salida, "Validación manual.csv")
    archivo_num = os.path.join(carpeta_salida, "Números utilizados.csv")

    # Exportar datos diarios a CSV
    if resumen["Datos_Validacion"]:
        keys = resumen["Datos_Validacion"][0].keys()
        # UTF-8 con BOM mantiene correctamente tildes y la letra eñe en Excel.
        with open(archivo_csv, 'w', newline='', encoding='utf-8-sig') as file:
            writer = csv.DictWriter(file, fieldnames=keys)
            writer.writeheader()
            writer.writerows(resumen["Datos_Validacion"])
        print(f"Exportado a CSV: {archivo_csv}")
            
        # Exportar a Excel en dos archivos separados
        try:
            df = pd.DataFrame(resumen["Datos_Validacion"])
            
            # 1. Excel de Demanda Diaria
            cols_demanda = ["Dia", "Stock_Inicial", "Llega_Pedido", "Demanda", "Stock_Final", "Ventas_Perdidas", "CTALM_Acumulado", "CVTAP_Acumulado", "CTF_Acumulado"]
            df_demanda = df[cols_demanda]
            excel_demanda = os.path.join(carpeta_salida, "Demanda diaria.xlsx")
            df_demanda.to_excel(excel_demanda, index=False)
            dar_formato_excel(excel_demanda)
            print(f"Exportado a Excel de Demanda: {excel_demanda}")
            
            # 2. Excel de Tiempo de Espera del Proveedor (Lead Time)
            cols_proveedor = ["Dia", "Stock_Inicial", "Stock_Final", "Emite_Pedido", "Lead_Time", "Llega_Pedido", "CTEP_Acumulado", "CTF_Acumulado"]
            df_proveedor = df[cols_proveedor]
            excel_proveedor = os.path.join(carpeta_salida, "Tiempo de entrega del proveedor.xlsx")
            df_proveedor.to_excel(excel_proveedor, index=False)
            dar_formato_excel(excel_proveedor)
            print(f"Exportado a Excel de Proveedor: {excel_proveedor}")

            # 3. Excel completo, equivalente a la planilla diaria de referencia
            excel_completo = os.path.join(carpeta_salida, "Simulación de inventario.xlsx")
            exportar_simulacion_completa(
                resumen,
                resultados_politicas,
                mejor_politica,
                filepath=excel_completo
            )
            
        except Exception as e:
            print(f"Error al generar los Excels de simulación: {e}")
            
    # Exportar números usados
    with open(archivo_num, 'w', newline='', encoding='utf-8-sig') as file:
        writer = csv.writer(file)
        writer.writerow(["Tipo", "Numero_Pseudoaleatorio"])
        for n in resumen["Numeros_Usados_Dem"]:
            writer.writerow(["Demanda", n])
        for n in resumen["Numeros_Usados_LT"]:
            writer.writerow(["Lead_Time", n])

def _dataframe_a_registros(filepath):
    """Lee un CSV y devuelve tipos JSON nativos para el generador del libro."""
    dataframe = pd.read_csv(filepath)
    return json.loads(dataframe.to_json(orient="records", force_ascii=False))


def _encontrar_node_bundled():
    """Prioriza el Node del runtime de Codex y usa el del sistema como respaldo."""
    bundled = (
        Path.home()
        / ".cache"
        / "codex-runtimes"
        / "codex-primary-runtime"
        / "dependencies"
        / "node"
        / "bin"
        / "node.exe"
    )
    if bundled.exists():
        return str(bundled)
    node_sistema = shutil.which("node")
    if node_sistema:
        return node_sistema
    raise RuntimeError("No se encontró Node.js para generar el libro Excel.")


def _generar_libro_unico(payload, carpeta_salida):
    archivo_json = os.path.join(carpeta_salida, "datos_libro_simulacion.json")
    cantidad_pares = payload["metadata"]["pares"]
    archivo_excel = os.path.join(
        carpeta_salida,
        f"Simulacion_Entrega_Pet_{cantidad_pares}_pares.xlsx",
    )
    builder = os.path.join(os.path.dirname(__file__), "generar_libro_simulacion.mjs")

    with open(archivo_json, "w", encoding="utf-8") as archivo:
        json.dump(payload, archivo, ensure_ascii=False)

    comando = [_encontrar_node_bundled(), builder, archivo_json, archivo_excel]
    preview_dir = os.environ.get("SIMULACION_PREVIEW_DIR")
    if preview_dir:
        comando.append(preview_dir)

    try:
        resultado = subprocess.run(
            comando,
            cwd=os.path.dirname(__file__),
            capture_output=True,
            text=True,
            encoding="utf-8",
            errors="replace",
        )
        if resultado.returncode != 0:
            detalle = resultado.stderr.strip() or resultado.stdout.strip()
            raise RuntimeError(f"No se pudo generar el libro Excel: {detalle}")
    finally:
        if os.path.exists(archivo_json):
            os.remove(archivo_json)
        # artifact-tool puede dejar un informe técnico junto al libro. No es
        # parte de la entrega: la carpeta de la corrida debe contener un Excel.
        informe_inspeccion = f"{archivo_excel}.inspect.ndjson"
        if os.path.exists(informe_inspeccion):
            os.remove(informe_inspeccion)

    return archivo_excel


def ejecutar_simulacion_consolidada():
    """Ejecuta la experimentación y crea un solo Excel con una hoja por par."""
    # Mantiene tildes, símbolos y emojis legibles incluso si la salida se
    # redirige desde PowerShell (donde Python suele heredar CP-1252).
    for flujo in (sys.stdout, sys.stderr):
        if hasattr(flujo, "reconfigure"):
            flujo.reconfigure(encoding="utf-8", errors="replace")

    import experimento_intervalo_confianza as experimento

    # La función imprime toda la información estadística en terminal y genera
    # los CSV intermedios utilizados por el libro consolidado.
    experimento.main(mostrar_operacion=False)

    carpeta_ic = experimento.CARPETA_SALIDA
    resumenes = _dataframe_a_registros(
        os.path.join(carpeta_ic, "resumen_politicas_ic95.csv")
    )
    evolucion = _dataframe_a_registros(
        os.path.join(carpeta_ic, "evolucion_intervalos_por_etapa.csv")
    )
    replicas = _dataframe_a_registros(
        os.path.join(carpeta_ic, "replicas_todos_los_pares.csv")
    )

    # El último estado registrado de cada par resume el descarte progresivo.
    estado_final = {}
    for fila in evolucion:
        estado_final[(int(fila["PEP"]), int(fila["TP"]))] = fila["Estado"]
    for fila in resumenes:
        clave = (int(fila["PEP"]), int(fila["TP"]))
        fila["Estado_Final"] = estado_final.get(clave, "DESCARTADA")

    simulador = SimuladorInventario("numeros_demanda.csv", "numeros_lead_time.csv")
    pares_detallados = []
    for indice, (pep, tp) in enumerate(experimento.PARES_A_EVALUAR, start=1):
        simulador.reiniciar_secuencias()
        detalle = simulador.simular_politica(
            pep,
            tp,
            TF=experimento.DIAS_POR_REPLICA,
            validacion_manual=True,
        )
        pares_detallados.append({
            "pep": pep,
            "tp": tp,
            "resumen": {
                "Demanda_Total": detalle["Demanda_Total"],
                "Ventas_Perdidas": detalle["Ventas_Perdidas"],
                "Pedidos_Realizados": detalle["Pedidos_Realizados"],
                "CTALM": detalle["CTALM"],
                "CVTAP": detalle["CVTAP"],
                "CTEP": detalle["CTEP"],
                "CTF": detalle["CTF"],
            },
            "dias": detalle["Datos_Validacion"],
        })

    carpeta_salida = crear_carpeta_corrida()
    payload = {
        "metadata": {
            "generado": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "replicas": experimento.REPLICAS,
            "dias": experimento.DIAS_POR_REPLICA,
            "pares": len(experimento.PARES_A_EVALUAR),
        },
        "resumen": resumenes,
        "evolucion": evolucion,
        "replicas": replicas,
        "pares": pares_detallados,
    }
    archivo_excel = _generar_libro_unico(payload, carpeta_salida)

    return archivo_excel


if __name__ == "__main__":
    ejecutar_simulacion_consolidada()
